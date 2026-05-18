"""
Report Builder Service
======================
Predefined reports with filters, consistent structure, and Excel export.

Each report returns::

    {
        "columns": [{"key": "...", "label": "..."}],
        "data": [{ ... }, ...],
        "summary": { ... }
    }

Reuses existing query logic from InventoryQueryService, ScrapAnalyticsService,
and StockValuationService where possible.
"""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Optional

from sqlalchemy.orm import Session

from ..models import (
    Customer,
    ProductionItem,
    ScrapRecord,
)
from ..models_v2 import (
    DispatchNote,
    DocumentStatus,
    GoodsReceiptNote,
    MaterialConsumptionV2,
    MaterialMaster,
    MovementType,
    QAStatus,
    StockLot,
    StockMovement,
    Vendor,
)
from .inventory_service import InventoryQueryService
from .stock_valuation_service import StockValuationService

ZERO = Decimal("0")
ZERO_2 = Decimal("0.00")
ZERO_3 = Decimal("0.000")


def _parse_date(value: Optional[str]) -> Optional[datetime]:
    """Parse an ISO date string to datetime, or return None."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def _safe_int(value) -> Optional[int]:
    """Convert to int or return None."""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


class ReportService:
    """Predefined reports with filters and Excel export."""

    # ------------------------------------------------------------------
    # 1. Stock Balance
    # ------------------------------------------------------------------
    @staticmethod
    def stock_balance(db: Session, filters: Optional[dict] = None) -> dict:
        """Current stock balance by material.

        Filters: material_id, location_id, qa_status
        """
        filters = filters or {}
        material_id = _safe_int(filters.get("material_id"))
        location_id = _safe_int(filters.get("location_id"))
        qa_status_str = filters.get("qa_status")
        qa_status = QAStatus(qa_status_str) if qa_status_str else None

        rows = InventoryQueryService.get_stock_summary(
            db,
            material_id=material_id,
            location_id=location_id,
            qa_status=qa_status,
        )

        total_qty = sum(r["total_weight_kg"] for r in rows)
        # Compute total value via valuation service
        total_value = 0.0
        for r in rows:
            val = StockValuationService.calculate_fifo_valuation(db, r["material_id"])
            r["total_value"] = val["total_value"]
            total_value += val["total_value"]

        columns = [
            {"key": "material_code", "label": "Material Code"},
            {"key": "material_name", "label": "Material Name"},
            {"key": "material_type", "label": "Type"},
            {"key": "grade", "label": "Grade"},
            {"key": "lot_count", "label": "Lots"},
            {"key": "total_weight_kg", "label": "Weight (KG)"},
            {"key": "total_weight_tons", "label": "Weight (Tons)"},
            {"key": "total_value", "label": "Value"},
        ]

        return {
            "columns": columns,
            "data": rows,
            "summary": {
                "total_qty_kg": round(total_qty, 3),
                "total_value": round(total_value, 2),
                "material_count": len(rows),
            },
        }

    # ------------------------------------------------------------------
    # 2. Stock Ledger
    # ------------------------------------------------------------------
    @staticmethod
    def stock_ledger(db: Session, filters: Optional[dict] = None) -> dict:
        """Chronological stock movement ledger.

        Filters: material_id, lot_id, date_from, date_to, movement_type
        """
        filters = filters or {}
        material_id = _safe_int(filters.get("material_id"))
        lot_id = _safe_int(filters.get("lot_id"))
        date_from = _parse_date(filters.get("date_from"))
        date_to = _parse_date(filters.get("date_to"))
        movement_type_str = filters.get("movement_type")

        query = (
            db.query(StockMovement, StockLot, MaterialMaster)
            .join(StockLot, StockMovement.stock_lot_id == StockLot.id)
            .join(MaterialMaster, StockLot.material_id == MaterialMaster.id)
        )

        if material_id:
            query = query.filter(StockLot.material_id == material_id)
        if lot_id:
            query = query.filter(StockMovement.stock_lot_id == lot_id)
        if date_from:
            query = query.filter(StockMovement.movement_date >= date_from)
        if date_to:
            query = query.filter(StockMovement.movement_date <= date_to)
        if movement_type_str:
            try:
                mt = MovementType(movement_type_str)
                query = query.filter(StockMovement.movement_type == mt)
            except ValueError:
                pass

        query = query.order_by(StockMovement.movement_date.asc(), StockMovement.id.asc())

        data = []
        for mov, lot, mat in query.all():
            mv_type = mov.movement_type.value if hasattr(mov.movement_type, "value") else str(mov.movement_type)
            data.append(
                {
                    "movement_number": mov.movement_number,
                    "movement_date": mov.movement_date.isoformat() if mov.movement_date else None,
                    "lot_number": lot.lot_number,
                    "material_code": mat.code,
                    "material_name": mat.name,
                    "movement_type": mv_type,
                    "weight_change_kg": float(mov.weight_change_kg),
                    "weight_before_kg": float(mov.weight_before_kg),
                    "weight_after_kg": float(mov.weight_after_kg),
                    "reference_type": mov.reference_type,
                    "reference_number": mov.reference_number,
                    "reason": mov.reason,
                    "valuation_rate": float(mov.valuation_rate) if mov.valuation_rate else None,
                    "stock_value_change": float(mov.stock_value_change) if mov.stock_value_change else None,
                }
            )

        columns = [
            {"key": "movement_number", "label": "Movement #"},
            {"key": "movement_date", "label": "Date"},
            {"key": "lot_number", "label": "Lot #"},
            {"key": "material_code", "label": "Material Code"},
            {"key": "material_name", "label": "Material Name"},
            {"key": "movement_type", "label": "Type"},
            {"key": "weight_change_kg", "label": "Change (KG)"},
            {"key": "weight_before_kg", "label": "Before (KG)"},
            {"key": "weight_after_kg", "label": "After (KG)"},
            {"key": "reference_number", "label": "Reference"},
            {"key": "reason", "label": "Reason"},
        ]

        return {
            "columns": columns,
            "data": data,
            "summary": {"total_movements": len(data)},
        }

    # ------------------------------------------------------------------
    # 3. Stock Aging
    # ------------------------------------------------------------------
    @staticmethod
    def stock_aging(db: Session, filters: Optional[dict] = None) -> dict:
        """FIFO aging report. Reuses InventoryQueryService.get_stock_aging_report.

        Filters: days_threshold (default 90)
        """
        filters = filters or {}
        days_threshold = _safe_int(filters.get("days_threshold")) or 90

        rows = InventoryQueryService.get_stock_aging_report(db, days_threshold=days_threshold)

        total_weight = sum(r["current_weight_kg"] for r in rows)
        old_stock_weight = sum(r["current_weight_kg"] for r in rows if r["is_old_stock"])

        # Serialize dates
        for r in rows:
            if r.get("received_date") and hasattr(r["received_date"], "isoformat"):
                r["received_date"] = r["received_date"].isoformat()

        columns = [
            {"key": "lot_number", "label": "Lot #"},
            {"key": "material_code", "label": "Material Code"},
            {"key": "material_name", "label": "Material Name"},
            {"key": "heat_number", "label": "Heat #"},
            {"key": "current_weight_kg", "label": "Weight (KG)"},
            {"key": "received_date", "label": "Received Date"},
            {"key": "age_days", "label": "Age (Days)"},
            {"key": "is_old_stock", "label": "Old Stock?"},
        ]

        return {
            "columns": columns,
            "data": rows,
            "summary": {
                "total_lots": len(rows),
                "total_weight_kg": round(total_weight, 3),
                "old_stock_weight_kg": round(old_stock_weight, 3),
                "days_threshold": days_threshold,
            },
        }

    # ------------------------------------------------------------------
    # 4. Material Consumption
    # ------------------------------------------------------------------
    @staticmethod
    def material_consumption(db: Session, filters: Optional[dict] = None) -> dict:
        """Material consumption summary by customer/period.

        Filters: date_from, date_to, customer_id, material_id
        """
        filters = filters or {}
        date_from = _parse_date(filters.get("date_from"))
        date_to = _parse_date(filters.get("date_to"))
        customer_id = _safe_int(filters.get("customer_id"))
        material_id = _safe_int(filters.get("material_id"))

        query = (
            db.query(
                MaterialConsumptionV2,
                StockLot,
                MaterialMaster,
            )
            .join(StockLot, MaterialConsumptionV2.stock_lot_id == StockLot.id)
            .join(MaterialMaster, StockLot.material_id == MaterialMaster.id)
        )

        if date_from:
            query = query.filter(MaterialConsumptionV2.consumed_at >= date_from)
        if date_to:
            query = query.filter(MaterialConsumptionV2.consumed_at <= date_to)
        if material_id:
            query = query.filter(StockLot.material_id == material_id)

        # customer_id filter via production item join
        if customer_id:
            from ..models_v2 import ProductionItemV2

            query = query.join(
                ProductionItemV2,
                MaterialConsumptionV2.production_item_id == ProductionItemV2.id,
            ).filter(ProductionItemV2.customer_id == customer_id)

        query = query.order_by(MaterialConsumptionV2.consumed_at.desc())

        data = []
        total_consumed_kg = 0.0
        total_value = 0.0

        for consumption, lot, mat in query.all():
            weight_kg = float(consumption.consumed_weight_kg)
            rate = float(lot.purchase_rate) if lot.purchase_rate else 0.0
            value = round(weight_kg * rate, 2)
            total_consumed_kg += weight_kg
            total_value += value

            data.append(
                {
                    "consumed_at": consumption.consumed_at.isoformat() if consumption.consumed_at else None,
                    "material_code": mat.code,
                    "material_name": mat.name,
                    "lot_number": lot.lot_number,
                    "consumed_weight_kg": weight_kg,
                    "rate_per_kg": rate,
                    "value": value,
                    "production_item_id": consumption.production_item_id,
                    "stage": consumption.stage,
                }
            )

        columns = [
            {"key": "consumed_at", "label": "Date"},
            {"key": "material_code", "label": "Material Code"},
            {"key": "material_name", "label": "Material"},
            {"key": "lot_number", "label": "Lot #"},
            {"key": "consumed_weight_kg", "label": "Consumed (KG)"},
            {"key": "rate_per_kg", "label": "Rate/KG"},
            {"key": "value", "label": "Value"},
            {"key": "stage", "label": "Stage"},
        ]

        return {
            "columns": columns,
            "data": data,
            "summary": {
                "total_consumed_kg": round(total_consumed_kg, 3),
                "total_value": round(total_value, 2),
                "record_count": len(data),
            },
        }

    # ------------------------------------------------------------------
    # 5. Scrap Analysis
    # ------------------------------------------------------------------
    @staticmethod
    def scrap_analysis(db: Session, filters: Optional[dict] = None) -> dict:
        """Scrap analysis report. Reuses ScrapAnalyticsService.

        Filters: date_from, date_to, reason_code, status
        """
        filters = filters or {}
        date_from = _parse_date(filters.get("date_from"))
        date_to = _parse_date(filters.get("date_to"))
        reason_code = filters.get("reason_code")
        status_filter = filters.get("status")

        query = db.query(ScrapRecord)

        if date_from:
            query = query.filter(ScrapRecord.created_at >= date_from)
        if date_to:
            query = query.filter(ScrapRecord.created_at <= date_to)
        if reason_code:
            query = query.filter(ScrapRecord.reason_code == reason_code)
        if status_filter:
            query = query.filter(ScrapRecord.status == status_filter)

        query = query.order_by(ScrapRecord.created_at.desc())

        records = query.all()
        data = []
        total_weight = 0.0
        total_value = 0.0

        for r in records:
            weight = float(r.weight_kg or 0)
            value = float(r.scrap_value or 0)
            total_weight += weight
            total_value += value
            data.append(
                {
                    "id": r.id,
                    "material_name": r.material_name,
                    "weight_kg": weight,
                    "quantity": r.quantity,
                    "reason_code": r.reason_code,
                    "status": r.status,
                    "scrap_value": value,
                    "dimensions": r.dimensions,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                }
            )

        # Scrap by reason breakdown
        by_reason: dict[str, float] = {}
        for r in records:
            reason = r.reason_code or "other"
            by_reason[reason] = by_reason.get(reason, 0) + float(r.weight_kg or 0)

        columns = [
            {"key": "material_name", "label": "Material"},
            {"key": "weight_kg", "label": "Weight (KG)"},
            {"key": "quantity", "label": "Qty"},
            {"key": "reason_code", "label": "Reason"},
            {"key": "status", "label": "Status"},
            {"key": "scrap_value", "label": "Scrap Value"},
            {"key": "created_at", "label": "Date"},
        ]

        return {
            "columns": columns,
            "data": data,
            "summary": {
                "total_weight_kg": round(total_weight, 3),
                "total_value": round(total_value, 2),
                "record_count": len(data),
                "by_reason": by_reason,
            },
        }

    # ------------------------------------------------------------------
    # 6. Production Progress
    # ------------------------------------------------------------------
    @staticmethod
    def production_progress(db: Session, filters: Optional[dict] = None) -> dict:
        """Production progress by customer/drawing.

        Filters: customer_id, stage, date_from, date_to
        """
        filters = filters or {}
        customer_id = _safe_int(filters.get("customer_id"))
        stage_filter = filters.get("stage")
        date_from = _parse_date(filters.get("date_from"))
        date_to = _parse_date(filters.get("date_to"))

        query = db.query(ProductionItem, Customer).join(Customer, ProductionItem.customer_id == Customer.id)

        if customer_id:
            query = query.filter(ProductionItem.customer_id == customer_id)
        if stage_filter:
            query = query.filter(ProductionItem.current_stage == stage_filter)
        if date_from:
            query = query.filter(ProductionItem.created_at >= date_from)
        if date_to:
            query = query.filter(ProductionItem.created_at <= date_to)

        query = query.order_by(ProductionItem.created_at.desc())

        items = query.all()
        data = []
        stage_counts: dict[str, int] = {}
        completed_count = 0

        for item, customer in items:
            current_stage = item.current_stage or "unknown"
            stage_counts[current_stage] = stage_counts.get(current_stage, 0) + 1
            if item.is_completed:
                completed_count += 1

            data.append(
                {
                    "id": item.id,
                    "customer_name": customer.name,
                    "item_code": item.item_code,
                    "item_name": item.item_name,
                    "section": item.section,
                    "quantity": item.quantity,
                    "current_stage": current_stage,
                    "is_completed": item.is_completed,
                    "created_at": item.created_at.isoformat() if item.created_at else None,
                }
            )

        total = len(data)
        completion_pct = round((completed_count / total * 100), 1) if total > 0 else 0.0

        columns = [
            {"key": "customer_name", "label": "Customer"},
            {"key": "item_code", "label": "Item Code"},
            {"key": "item_name", "label": "Item Name"},
            {"key": "section", "label": "Section"},
            {"key": "quantity", "label": "Qty"},
            {"key": "current_stage", "label": "Stage"},
            {"key": "is_completed", "label": "Completed"},
            {"key": "created_at", "label": "Created"},
        ]

        return {
            "columns": columns,
            "data": data,
            "summary": {
                "total_items": total,
                "completed": completed_count,
                "completion_pct": completion_pct,
                "stage_counts": stage_counts,
            },
        }

    # ------------------------------------------------------------------
    # 7. GRN Register
    # ------------------------------------------------------------------
    @staticmethod
    def grn_register(db: Session, filters: Optional[dict] = None) -> dict:
        """GRN register - all goods receipts.

        Filters: date_from, date_to, vendor_id, status
        """
        filters = filters or {}
        date_from = _parse_date(filters.get("date_from"))
        date_to = _parse_date(filters.get("date_to"))
        vendor_id = _safe_int(filters.get("vendor_id"))
        status_filter = filters.get("status")

        query = db.query(GoodsReceiptNote, Vendor).join(Vendor, GoodsReceiptNote.vendor_id == Vendor.id)

        if date_from:
            query = query.filter(GoodsReceiptNote.created_at >= date_from)
        if date_to:
            query = query.filter(GoodsReceiptNote.created_at <= date_to)
        if vendor_id:
            query = query.filter(GoodsReceiptNote.vendor_id == vendor_id)
        if status_filter:
            try:
                ds = DocumentStatus(status_filter)
                query = query.filter(GoodsReceiptNote.status == ds)
            except ValueError:
                pass

        query = query.order_by(GoodsReceiptNote.created_at.desc())

        data = []
        for grn, vendor in query.all():
            grn_status = grn.status.value if hasattr(grn.status, "value") else str(grn.status)
            data.append(
                {
                    "grn_number": grn.grn_number,
                    "vendor_name": vendor.name,
                    "vendor_code": vendor.code,
                    "vehicle_number": grn.vehicle_number,
                    "status": grn_status,
                    "net_weight_kg": float(grn.net_weight_kg) if grn.net_weight_kg else None,
                    "vendor_invoice_number": grn.vendor_invoice_number,
                    "gate_entry_time": grn.gate_entry_time.isoformat() if grn.gate_entry_time else None,
                    "created_at": grn.created_at.isoformat() if grn.created_at else None,
                }
            )

        columns = [
            {"key": "grn_number", "label": "GRN #"},
            {"key": "vendor_name", "label": "Vendor"},
            {"key": "vendor_code", "label": "Vendor Code"},
            {"key": "vehicle_number", "label": "Vehicle"},
            {"key": "status", "label": "Status"},
            {"key": "net_weight_kg", "label": "Net Weight (KG)"},
            {"key": "vendor_invoice_number", "label": "Invoice #"},
            {"key": "gate_entry_time", "label": "Gate Entry"},
            {"key": "created_at", "label": "Created"},
        ]

        return {
            "columns": columns,
            "data": data,
            "summary": {"total_grns": len(data)},
        }

    # ------------------------------------------------------------------
    # 8. Dispatch Register
    # ------------------------------------------------------------------
    @staticmethod
    def dispatch_register(db: Session, filters: Optional[dict] = None) -> dict:
        """Dispatch register - all dispatches.

        Filters: date_from, date_to, customer_id, status
        """
        filters = filters or {}
        date_from = _parse_date(filters.get("date_from"))
        date_to = _parse_date(filters.get("date_to"))
        customer_id = _safe_int(filters.get("customer_id"))
        status_filter = filters.get("status")

        query = db.query(DispatchNote, Customer).join(Customer, DispatchNote.customer_id == Customer.id)

        if date_from:
            query = query.filter(DispatchNote.created_at >= date_from)
        if date_to:
            query = query.filter(DispatchNote.created_at <= date_to)
        if customer_id:
            query = query.filter(DispatchNote.customer_id == customer_id)
        if status_filter:
            try:
                ds = DocumentStatus(status_filter)
                query = query.filter(DispatchNote.status == ds)
            except ValueError:
                pass

        query = query.order_by(DispatchNote.created_at.desc())

        data = []
        for dispatch, customer in query.all():
            d_status = dispatch.status.value if hasattr(dispatch.status, "value") else str(dispatch.status)
            data.append(
                {
                    "dispatch_number": dispatch.dispatch_number,
                    "customer_name": customer.name,
                    "vehicle_number": dispatch.vehicle_number,
                    "transporter": dispatch.transporter,
                    "status": d_status,
                    "net_weight_kg": float(dispatch.net_weight_kg) if dispatch.net_weight_kg else None,
                    "dispatched_at": dispatch.dispatched_at.isoformat() if dispatch.dispatched_at else None,
                    "created_at": dispatch.created_at.isoformat() if dispatch.created_at else None,
                }
            )

        columns = [
            {"key": "dispatch_number", "label": "Dispatch #"},
            {"key": "customer_name", "label": "Customer"},
            {"key": "vehicle_number", "label": "Vehicle"},
            {"key": "transporter", "label": "Transporter"},
            {"key": "status", "label": "Status"},
            {"key": "net_weight_kg", "label": "Net Weight (KG)"},
            {"key": "dispatched_at", "label": "Dispatched"},
            {"key": "created_at", "label": "Created"},
        ]

        return {
            "columns": columns,
            "data": data,
            "summary": {"total_dispatches": len(data)},
        }

    # ------------------------------------------------------------------
    # Excel Export
    # ------------------------------------------------------------------
    @staticmethod
    def export_to_excel(report_data: dict, report_name: str) -> BytesIO:
        """Convert any report data dict to an Excel file (BytesIO).

        Uses openpyxl. The columns list drives the header row and cell order.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = report_name[:31]  # Excel limits sheet name to 31 chars

        columns = report_data.get("columns", [])
        data = report_data.get("data", [])

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.get("label", col_def.get("key", "")))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_def in enumerate(columns, start=1):
                key = col_def["key"]
                value = row_data.get(key)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns (approximate)
        for col_idx, col_def in enumerate(columns, start=1):
            label = col_def.get("label", col_def.get("key", ""))
            max_len = len(str(label))
            for row_data in data[:100]:  # sample first 100 rows
                val = row_data.get(col_def["key"])
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# ======================================================================
# Report Registry
# ======================================================================

REPORT_REGISTRY: dict[str, callable] = {
    "stock-balance": ReportService.stock_balance,
    "stock-ledger": ReportService.stock_ledger,
    "stock-aging": ReportService.stock_aging,
    "material-consumption": ReportService.material_consumption,
    "scrap-analysis": ReportService.scrap_analysis,
    "production-progress": ReportService.production_progress,
    "grn-register": ReportService.grn_register,
    "dispatch-register": ReportService.dispatch_register,
}

REPORT_DESCRIPTIONS: dict[str, str] = {
    "stock-balance": "Current stock balance by material with valuation",
    "stock-ledger": "Chronological stock movement ledger",
    "stock-aging": "FIFO aging report for stock management",
    "material-consumption": "Material consumption summary by customer/period",
    "scrap-analysis": "Scrap analysis by reason, status, and period",
    "production-progress": "Production progress by customer and stage",
    "grn-register": "Goods Receipt Note register",
    "dispatch-register": "Dispatch register",
}
