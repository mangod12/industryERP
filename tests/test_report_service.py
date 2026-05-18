"""
Tests for ReportService and /api/v2/reports/ endpoints.

Covers:
- Each report returns correct {columns, data, summary} structure
- Filters work (date range, material_id, etc.)
- Empty data returns empty list (not error)
- Excel export returns valid xlsx (check header bytes)
- List reports returns all 8
- Invalid report name returns 404
- Router integration tests
"""

import sys
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

# Ensure imports work
_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from backend_core.app.models import (
    Customer,
    ScrapRecord,
)
from backend_core.app.models_v2 import (
    DispatchNote,
    DocumentStatus,
    GoodsReceiptNote,
    MaterialConsumptionV2,
    MaterialMaster,
    MaterialType,
    MovementType,
    ProductionItemV2,
    QAStatus,
    StockLot,
    StockMovement,
    Vendor,
)
from backend_core.app.services.report_service import (
    REPORT_DESCRIPTIONS,
    REPORT_REGISTRY,
    ReportService,
)
from tests.conftest import (
    create_test_customer,
    create_test_production_item,
    create_test_user,
)

# ======================================================================
# Helpers
# ======================================================================


def _assert_report_structure(result: dict):
    """Assert the report has the standard {columns, data, summary} shape."""
    assert "columns" in result, "Report missing 'columns'"
    assert "data" in result, "Report missing 'data'"
    assert "summary" in result, "Report missing 'summary'"
    assert isinstance(result["columns"], list)
    assert isinstance(result["data"], list)
    assert isinstance(result["summary"], dict)
    # Each column must have key and label
    for col in result["columns"]:
        assert "key" in col
        assert "label" in col


def _create_material(db, code="MAT-RPT-001", name="Test Steel Plate"):
    """Create a MaterialMaster row."""
    mat = MaterialMaster(
        code=code,
        name=name,
        material_type=MaterialType.PLATE,
        grade="IS2062",
    )
    db.add(mat)
    db.commit()
    db.refresh(mat)
    return mat


def _create_vendor(db, code="V-RPT-001", name="Test Vendor"):
    """Create a Vendor row."""
    vendor = Vendor(code=code, name=name)
    db.add(vendor)
    db.commit()
    db.refresh(vendor)
    return vendor


def _create_stock_lot_with_movement(db, material, vendor, user, weight=Decimal("500.000")):
    """Create a stock lot and an inward movement for testing."""
    lot = StockLot(
        lot_number=f"LOT-RPT-{material.id}-{int(datetime.utcnow().timestamp() * 1000)}",
        material_id=material.id,
        vendor_id=vendor.id,
        gross_weight_kg=weight + Decimal("10"),
        tare_weight_kg=Decimal("10"),
        net_weight_kg=weight,
        current_weight_kg=weight,
        received_date=datetime.utcnow(),
        qa_status=QAStatus.APPROVED,
        purchase_rate=Decimal("45.00"),
    )
    db.add(lot)
    db.flush()

    movement = StockMovement(
        movement_number=f"MOV-RPT-{lot.id}-{int(datetime.utcnow().timestamp() * 1000)}",
        stock_lot_id=lot.id,
        movement_type=MovementType.INWARD_PURCHASE,
        weight_change_kg=weight,
        weight_before_kg=Decimal("0"),
        weight_after_kg=weight,
        quantity_change=Decimal("1"),
        reference_type="grn",
        created_by=user.id,
        movement_date=datetime.utcnow(),
    )
    db.add(movement)
    db.commit()
    db.refresh(lot)
    return lot, movement


# ======================================================================
# Unit Tests — ReportService
# ======================================================================


class TestReportRegistry:
    """Tests for the report registry."""

    def test_registry_has_8_reports(self):
        assert len(REPORT_REGISTRY) == 8

    def test_all_reports_have_descriptions(self):
        for name in REPORT_REGISTRY:
            assert name in REPORT_DESCRIPTIONS

    def test_registry_keys(self):
        expected = {
            "stock-balance",
            "stock-ledger",
            "stock-aging",
            "material-consumption",
            "scrap-analysis",
            "production-progress",
            "grn-register",
            "dispatch-register",
        }
        assert set(REPORT_REGISTRY.keys()) == expected


class TestStockBalanceReport:
    """Tests for the stock-balance report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.stock_balance(db)
        _assert_report_structure(result)
        assert result["data"] == []
        assert result["summary"]["total_qty_kg"] == 0

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db)
        vendor = _create_vendor(db)
        _create_stock_lot_with_movement(db, mat, vendor, user)

        result = ReportService.stock_balance(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1
        assert result["summary"]["total_qty_kg"] > 0

    def test_filter_by_material(self, db):
        user = create_test_user(db, role="Boss")
        mat1 = _create_material(db, code="MAT-A1", name="Steel A")
        mat2 = _create_material(db, code="MAT-B1", name="Steel B")
        vendor = _create_vendor(db)
        _create_stock_lot_with_movement(db, mat1, vendor, user)
        _create_stock_lot_with_movement(db, mat2, vendor, user)

        result = ReportService.stock_balance(db, filters={"material_id": str(mat1.id)})
        _assert_report_structure(result)
        assert len(result["data"]) == 1
        assert result["data"][0]["material_code"] == "MAT-A1"


class TestStockLedgerReport:
    """Tests for the stock-ledger report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.stock_ledger(db)
        _assert_report_structure(result)
        assert result["data"] == []

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db, code="MAT-LED-001")
        vendor = _create_vendor(db, code="V-LED-001")
        _create_stock_lot_with_movement(db, mat, vendor, user)

        result = ReportService.stock_ledger(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1
        assert result["summary"]["total_movements"] >= 1

    def test_filter_by_movement_type(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db, code="MAT-LED-002")
        vendor = _create_vendor(db, code="V-LED-002")
        _create_stock_lot_with_movement(db, mat, vendor, user)

        result = ReportService.stock_ledger(db, filters={"movement_type": "inward_purchase"})
        _assert_report_structure(result)
        assert len(result["data"]) >= 1

        result_empty = ReportService.stock_ledger(db, filters={"movement_type": "outward_sale"})
        assert result_empty["data"] == []

    def test_filter_by_date_range(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db, code="MAT-LED-003")
        vendor = _create_vendor(db, code="V-LED-003")
        _create_stock_lot_with_movement(db, mat, vendor, user)

        future = (datetime.utcnow() + timedelta(days=10)).isoformat()
        result = ReportService.stock_ledger(db, filters={"date_from": future})
        assert result["data"] == []


class TestStockAgingReport:
    """Tests for the stock-aging report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.stock_aging(db)
        _assert_report_structure(result)
        assert result["data"] == []
        assert result["summary"]["total_lots"] == 0

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db, code="MAT-AGE-001")
        vendor = _create_vendor(db, code="V-AGE-001")
        _create_stock_lot_with_movement(db, mat, vendor, user)

        result = ReportService.stock_aging(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1


class TestMaterialConsumptionReport:
    """Tests for the material-consumption report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.material_consumption(db)
        _assert_report_structure(result)
        assert result["data"] == []
        assert result["summary"]["total_consumed_kg"] == 0

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db, code="MAT-CON-001")
        vendor = _create_vendor(db, code="V-CON-001")
        lot, _ = _create_stock_lot_with_movement(db, mat, vendor, user)

        customer = Customer(
            name="Consumption Customer",
            project_details="test",
            email="cons@test.com",
            phone="1234567890",
        )
        db.add(customer)
        db.flush()

        prod_item = ProductionItemV2(
            customer_id=customer.id,
            item_code="PI-CON-001",
            item_name="Test Production Item",
        )
        db.add(prod_item)
        db.flush()

        consumption = MaterialConsumptionV2(
            production_item_id=prod_item.id,
            stock_lot_id=lot.id,
            consumed_weight_kg=Decimal("50.000"),
            consumed_by=user.id,
            consumed_at=datetime.utcnow(),
        )
        db.add(consumption)
        db.commit()

        result = ReportService.material_consumption(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1
        assert result["summary"]["total_consumed_kg"] > 0


class TestScrapAnalysisReport:
    """Tests for the scrap-analysis report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.scrap_analysis(db)
        _assert_report_structure(result)
        assert result["data"] == []

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        scrap = ScrapRecord(
            material_name="Test Plate",
            weight_kg=25.5,
            reason_code="cutting_waste",
            status="pending",
            created_by=user.id,
            created_at=datetime.utcnow(),
        )
        db.add(scrap)
        db.commit()

        result = ReportService.scrap_analysis(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1
        assert result["summary"]["total_weight_kg"] > 0

    def test_filter_by_reason(self, db):
        user = create_test_user(db, role="Boss")
        s1 = ScrapRecord(
            material_name="Plate A",
            weight_kg=10.0,
            reason_code="cutting_waste",
            status="pending",
            created_by=user.id,
            created_at=datetime.utcnow(),
        )
        s2 = ScrapRecord(
            material_name="Plate B",
            weight_kg=20.0,
            reason_code="defect",
            status="pending",
            created_by=user.id,
            created_at=datetime.utcnow(),
        )
        db.add_all([s1, s2])
        db.commit()

        result = ReportService.scrap_analysis(db, filters={"reason_code": "cutting_waste"})
        assert all(r["reason_code"] == "cutting_waste" for r in result["data"])


class TestProductionProgressReport:
    """Tests for the production-progress report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.production_progress(db)
        _assert_report_structure(result)
        assert result["data"] == []

    def test_with_data(self, db):
        customer = create_test_customer(db, name="Progress Customer")
        create_test_production_item(db, customer.id, item_code="P1", current_stage="fabrication")
        create_test_production_item(db, customer.id, item_code="P2", current_stage="painting")

        result = ReportService.production_progress(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 2
        assert "stage_counts" in result["summary"]

    def test_filter_by_stage(self, db):
        customer = create_test_customer(db, name="Stage Filter Customer")
        create_test_production_item(db, customer.id, item_code="SF1", current_stage="fabrication")
        create_test_production_item(db, customer.id, item_code="SF2", current_stage="dispatch")

        result = ReportService.production_progress(db, filters={"stage": "fabrication"})
        assert all(r["current_stage"] == "fabrication" for r in result["data"])


class TestGRNRegisterReport:
    """Tests for the grn-register report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.grn_register(db)
        _assert_report_structure(result)
        assert result["data"] == []

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        vendor = _create_vendor(db, code="V-GRN-001")
        grn = GoodsReceiptNote(
            grn_number="GRN-RPT-001",
            vendor_id=vendor.id,
            status=DocumentStatus.DRAFT,
            created_by=user.id,
            gate_entry_time=datetime.utcnow(),
        )
        db.add(grn)
        db.commit()

        result = ReportService.grn_register(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1
        assert result["summary"]["total_grns"] >= 1


class TestDispatchRegisterReport:
    """Tests for the dispatch-register report."""

    def test_empty_returns_structure(self, db):
        result = ReportService.dispatch_register(db)
        _assert_report_structure(result)
        assert result["data"] == []

    def test_with_data(self, db):
        user = create_test_user(db, role="Boss")
        customer = create_test_customer(db, name="Dispatch Customer")
        dispatch = DispatchNote(
            dispatch_number="DN-RPT-001",
            customer_id=customer.id,
            status=DocumentStatus.DRAFT,
            created_by=user.id,
        )
        db.add(dispatch)
        db.commit()

        result = ReportService.dispatch_register(db)
        _assert_report_structure(result)
        assert len(result["data"]) >= 1
        assert result["summary"]["total_dispatches"] >= 1


class TestExcelExport:
    """Tests for Excel export."""

    def test_export_empty_report(self, db):
        result = ReportService.stock_balance(db)
        excel = ReportService.export_to_excel(result, "stock-balance")
        assert excel is not None
        # Check xlsx magic bytes (PK zip header)
        content = excel.read()
        assert content[:2] == b"PK"
        assert len(content) > 100

    def test_export_with_data(self, db):
        user = create_test_user(db, role="Boss")
        mat = _create_material(db, code="MAT-XLS-001")
        vendor = _create_vendor(db, code="V-XLS-001")
        _create_stock_lot_with_movement(db, mat, vendor, user)

        result = ReportService.stock_balance(db)
        excel = ReportService.export_to_excel(result, "stock-balance")
        content = excel.read()
        assert content[:2] == b"PK"
        assert len(content) > 200

    def test_export_preserves_columns(self, db):
        """Verify the workbook has header row matching column labels."""

        from openpyxl import load_workbook

        result = ReportService.stock_aging(db)
        excel = ReportService.export_to_excel(result, "stock-aging")
        wb = load_workbook(excel)
        ws = wb.active

        expected_labels = [c["label"] for c in result["columns"]]
        actual_labels = [ws.cell(row=1, column=i + 1).value for i in range(len(expected_labels))]
        assert actual_labels == expected_labels


# ======================================================================
# Integration Tests — Router
# ======================================================================


class TestReportsRouter:
    """Tests for the /api/v2/reports/ endpoints."""

    def test_list_reports(self, boss_client):
        resp = boss_client.get("/api/v2/reports/")
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is True
        reports = body["data"]["reports"]
        assert len(reports) == 8
        report_names = {r["name"] for r in reports}
        assert "stock-balance" in report_names
        assert "dispatch-register" in report_names

    def test_invalid_report_404(self, boss_client):
        resp = boss_client.get("/api/v2/reports/nonexistent-report")
        assert resp.status_code == 404

    def test_get_stock_balance(self, boss_client):
        resp = boss_client.get("/api/v2/reports/stock-balance")
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is True
        _assert_report_structure(body["data"])

    def test_get_stock_ledger(self, boss_client):
        resp = boss_client.get("/api/v2/reports/stock-ledger")
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is True
        _assert_report_structure(body["data"])

    def test_get_stock_aging(self, boss_client):
        resp = boss_client.get("/api/v2/reports/stock-aging")
        assert resp.status_code == 200
        _assert_report_structure(resp.json()["data"])

    def test_get_material_consumption(self, boss_client):
        resp = boss_client.get("/api/v2/reports/material-consumption")
        assert resp.status_code == 200
        _assert_report_structure(resp.json()["data"])

    def test_get_scrap_analysis(self, boss_client):
        resp = boss_client.get("/api/v2/reports/scrap-analysis")
        assert resp.status_code == 200
        _assert_report_structure(resp.json()["data"])

    def test_get_production_progress(self, boss_client):
        resp = boss_client.get("/api/v2/reports/production-progress")
        assert resp.status_code == 200
        _assert_report_structure(resp.json()["data"])

    def test_get_grn_register(self, boss_client):
        resp = boss_client.get("/api/v2/reports/grn-register")
        assert resp.status_code == 200
        _assert_report_structure(resp.json()["data"])

    def test_get_dispatch_register(self, boss_client):
        resp = boss_client.get("/api/v2/reports/dispatch-register")
        assert resp.status_code == 200
        _assert_report_structure(resp.json()["data"])

    def test_export_stock_balance_excel(self, boss_client):
        resp = boss_client.get("/api/v2/reports/stock-balance/export")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")
        assert resp.content[:2] == b"PK"

    def test_export_invalid_report_404(self, boss_client):
        resp = boss_client.get("/api/v2/reports/nonexistent/export")
        assert resp.status_code == 404

    def test_query_params_passed_as_filters(self, boss_client):
        resp = boss_client.get("/api/v2/reports/stock-ledger?movement_type=outward_sale")
        assert resp.status_code == 200
        body = resp.json()
        _assert_report_structure(body["data"])
        # With no data, should be empty
        assert body["data"]["data"] == []

    def test_unauthenticated_returns_401(self, client):
        resp = client.get("/api/v2/reports/")
        assert resp.status_code == 401
